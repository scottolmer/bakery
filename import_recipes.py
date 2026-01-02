"""
Import recipes from Excel spreadsheet to database
"""
import openpyxl
from app import app
from models import db, Recipe, Ingredient, RecipeIngredient

# Common bread recipes and their typical structure
RECIPE_SHEETS = [
    'Italian', 'Multigrain', 'Rustic White', 'Baguette', 'Pain dMie',
    'New Miche', 'Brioche', 'Schiacciata', 'Dinkel', 'Fino', 'Croissant',
    'Two-Day Italian', 'Two-Day Multigrain', 'Two-Day Baguette',
    'Light Rye', 'Dark Rye', 'Ciabatta', 'Naan', 'Miche', 'Focaccia',
    'Stollen', 'Pumpkin Miche', 'Hot Cross Buns', 'Brotchen',
    'Chocolate Croissant', 'Irish Soda Bread'
]

STARTER_SHEETS = ['Levain', 'Biga', 'Poolish', 'Emmy(starter)', 'Itl Levain']
SOAKER_SHEETS = ['RW Soaker', '7 Grain Soaker', 'Dinkel Soaker']


def analyze_recipe_sheet(ws, sheet_name):
    """
    Analyze a recipe sheet to extract structure
    """
    print(f"\n=== Analyzing {sheet_name} ===")

    # Look for key patterns
    batch_weight = None
    loaf_weight = None
    ingredients = []

    # Scan first 50 rows for recipe data
    for row_idx in range(1, 51):
        for col_idx in range(1, 10):  # Check first 10 columns
            cell = ws.cell(row_idx, col_idx)

            if cell.value:
                cell_str = str(cell.value).strip()

                # Look for batch weight
                if 'Batch Weight' in cell_str or 'batch weight' in cell_str:
                    # Try next cell
                    next_cell = ws.cell(row_idx, col_idx + 1)
                    if next_cell.value and isinstance(next_cell.value, (int, float)):
                        batch_weight = float(next_cell.value)
                        print(f"  Found batch weight: {batch_weight}g at {cell.coordinate}")

                # Look for loaf weight (like "1000 grams" or "400 grams")
                if 'grams' in cell_str.lower() and col_idx == 1:
                    import re
                    match = re.search(r'(\d+)\s*g', cell_str)
                    if match:
                        potential_loaf = int(match.group(1))
                        if 100 <= potential_loaf <= 3000:  # Reasonable loaf weight range
                            loaf_weight = potential_loaf
                            print(f"  Found loaf weight: {loaf_weight}g at {cell.coordinate}")

                # Look for ingredient names in column A or D
                if col_idx in [1, 4]:
                    # Common ingredient keywords
                    ingredient_keywords = [
                        'flour', 'water', 'salt', 'yeast', 'levain', 'poolish',
                        'biga', 'oil', 'sugar', 'milk', 'butter', 'egg', 'soaker',
                        'Red Rose', 'Whole Wheat', 'Rye'
                    ]

                    for keyword in ingredient_keywords:
                        if keyword.lower() in cell_str.lower():
                            # Get value from next column
                            value_cell = ws.cell(row_idx, col_idx + 1)
                            if value_cell.value and isinstance(value_cell.value, (int, float)):
                                ingredients.append({
                                    'name': cell_str,
                                    'amount': float(value_cell.value),
                                    'row': row_idx,
                                    'col': col_idx
                                })
                                print(f"  Found ingredient: {cell_str} = {value_cell.value} at {cell.coordinate}")
                            break

    return {
        'name': sheet_name,
        'batch_weight': batch_weight,
        'loaf_weight': loaf_weight,
        'ingredients': ingredients
    }


def import_recipes_from_excel(filename='Bread Formulas 2024.xlsx'):
    """
    Import all recipes from Excel file
    """
    print(f"Loading {filename}...")
    wb = openpyxl.load_workbook(filename, data_only=True)

    with app.app_context():
        # First, ensure we have all common ingredients
        create_common_ingredients()

        # Import starters
        print("\n" + "="*60)
        print("IMPORTING STARTERS")
        print("="*60)
        for sheet_name in STARTER_SHEETS:
            if sheet_name in wb.sheetnames:
                import_recipe(wb[sheet_name], sheet_name, 'starter')

        # Import soakers
        print("\n" + "="*60)
        print("IMPORTING SOAKERS")
        print("="*60)
        for sheet_name in SOAKER_SHEETS:
            if sheet_name in wb.sheetnames:
                import_recipe(wb[sheet_name], sheet_name, 'soaker')

        # Import breads
        print("\n" + "="*60)
        print("IMPORTING BREAD RECIPES")
        print("="*60)
        for sheet_name in RECIPE_SHEETS:
            if sheet_name in wb.sheetnames:
                import_recipe(wb[sheet_name], sheet_name, 'bread')

    wb.close()
    print("\n" + "="*60)
    print("IMPORT COMPLETE!")
    print("="*60)


def create_common_ingredients():
    """
    Create common ingredients if they don't exist
    """
    ingredients_data = [
        ('Red Rose Flour', 'flour'),
        ('Whole Wheat Flour', 'flour'),
        ('Spelt Flour', 'flour'),
        ('Rye Flour', 'flour'),
        ('Water', 'water'),
        ('Salt', 'salt'),
        ('Instant Yeast', 'yeast'),
        ('Dry Yeast', 'yeast'),
        ('Gold Yeast', 'yeast'),
        ('Cake Yeast', 'yeast'),
        ('Levain', 'starter'),
        ('Poolish', 'starter'),
        ('Biga', 'starter'),
        ('Rye Levain', 'starter'),
        ('Red Rose Levain', 'starter'),
        ('Levain Discard', 'starter'),
        ('Itl Levain', 'starter'),
        ('Olive Oil', 'oil'),
        ('Butter', 'dairy'),
        ('Whole Milk', 'dairy'),
        ('Milk Powder', 'dairy'),
        ('Sugar', 'sugar'),
        ('Honey', 'sugar'),
        ('Barley Malt', 'sugar'),
        ('Eggs', 'dairy'),
        ('Whole Egg', 'dairy'),
        ('7 Grain Soaker', 'soaker'),
        ('RW Soaker', 'soaker'),
        ('Dinkel Soaker', 'soaker'),
        ('MG Soaker', 'soaker'),
        ('Spice Mix', 'other'),
        ('Polenta', 'other'),
    ]

    for name, category in ingredients_data:
        if not Ingredient.query.filter_by(name=name).first():
            ingredient = Ingredient(name=name, category=category)
            db.session.add(ingredient)
            print(f"Created ingredient: {name}")

    db.session.commit()
    print(f"\nIngredients initialized!")


def import_recipe(ws, sheet_name, recipe_type):
    """
    Import a single recipe from a worksheet
    """
    print(f"\nImporting {sheet_name}...")

    # Check if recipe already exists
    existing = Recipe.query.filter_by(name=sheet_name).first()
    if existing:
        print(f"  [SKIP] Recipe '{sheet_name}' already exists, skipping...")
        return

    # Analyze the sheet
    data = analyze_recipe_sheet(ws, sheet_name)

    # Use defaults if we couldn't extract values
    batch_weight = data['batch_weight'] or 1000
    loaf_weight = data['loaf_weight'] or 1000

    # Create recipe
    recipe = Recipe(
        name=sheet_name,
        recipe_type=recipe_type,
        base_batch_weight=batch_weight,
        loaf_weight=loaf_weight,
        is_active=True
    )
    db.session.add(recipe)
    db.session.flush()  # Get the ID

    print(f"  [OK] Created recipe: {sheet_name}")
    print(f"    - Type: {recipe_type}")
    print(f"    - Batch Weight: {batch_weight}g")
    print(f"    - Loaf Weight: {loaf_weight}g")

    # Add ingredients
    if data['ingredients']:
        print(f"  Adding {len(data['ingredients'])} ingredients:")
        for idx, ing_data in enumerate(data['ingredients'][:10]):  # Limit to first 10
            # Try to find matching ingredient
            ingredient = find_or_create_ingredient(ing_data['name'])

            if ingredient:
                # Assume baker's percentage for now
                recipe_ing = RecipeIngredient(
                    recipe_id=recipe.id,
                    ingredient_id=ingredient.id,
                    percentage=ing_data['amount'],
                    is_percentage=True,
                    order=idx
                )
                db.session.add(recipe_ing)
                print(f"    - {ingredient.name}: {ing_data['amount']}%")

    db.session.commit()
    print(f"  [OK] Recipe '{sheet_name}' imported successfully!")


def find_or_create_ingredient(name):
    """
    Find ingredient by name or create if not found
    """
    # Clean up the name
    name = name.strip()

    # Try exact match first
    ingredient = Ingredient.query.filter_by(name=name).first()
    if ingredient:
        return ingredient

    # Try partial matches
    for ing in Ingredient.query.all():
        if ing.name.lower() in name.lower() or name.lower() in ing.name.lower():
            return ing

    # Create new ingredient
    print(f"    Creating new ingredient: {name}")
    ingredient = Ingredient(name=name, category='other')
    db.session.add(ingredient)
    db.session.flush()
    return ingredient


if __name__ == '__main__':
    import_recipes_from_excel()
