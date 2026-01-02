"""
Utilities for extracting data from Excel spreadsheets
"""
import openpyxl
from typing import Dict, List, Tuple, Any
import re


class ExcelParser:
    """Parse bread formula and order Excel files"""

    def __init__(self, formulas_file: str, orders_file: str):
        self.formulas_file = formulas_file
        self.orders_file = orders_file
        self.formulas_wb = None
        self.orders_wb = None

    def load_workbooks(self):
        """Load Excel workbooks"""
        print(f"Loading {self.formulas_file}...")
        self.formulas_wb = openpyxl.load_workbook(self.formulas_file, data_only=False)
        print(f"Loading {self.orders_file}...")
        self.orders_wb = openpyxl.load_workbook(self.orders_file, data_only=False)

    def get_recipe_sheets(self) -> List[str]:
        """Get list of recipe sheet names"""
        # Common recipe sheets based on what we saw
        recipe_sheets = [
            'Italian', 'Multigrain', 'Rustic White', 'Baguette',
            'Pain dMie', 'New Miche', 'Brioche', 'Schiacciata',
            'Dinkel', 'Fino', 'Croissant', 'Chocolate Croissant',
            'Two-Day Italian', 'Two-Day Multigrain', 'Two-Day Baguette',
            'Light Rye', 'Dark Rye', 'Ciabatta', 'Naan', 'Miche',
            'Focaccia', 'Stollen', 'Pumpkin Miche', 'Hot Cross Buns',
            'Brotchen', 'Irish Soda Bread'
        ]
        # Filter to only sheets that exist
        return [s for s in recipe_sheets if s in self.formulas_wb.sheetnames]

    def get_starter_sheets(self) -> List[str]:
        """Get list of starter/preferment sheets"""
        starter_sheets = [
            'Levain', 'Biga', 'Poolish', 'Emmy(starter)',
            'Itl Levain', 'Starters'
        ]
        return [s for s in starter_sheets if s in self.formulas_wb.sheetnames]

    def get_soaker_sheets(self) -> List[str]:
        """Get list of soaker sheets"""
        soaker_sheets = [
            'RW Soaker', '7 Grain Soaker', 'Dinkel Soaker', 'Soaks'
        ]
        return [s for s in soaker_sheets if s in self.formulas_wb.sheetnames]

    def extract_recipe_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Extract recipe data from a sheet
        Returns dict with recipe name, ingredients, and formulas
        """
        if self.formulas_wb is None:
            self.load_workbooks()

        ws = self.formulas_wb[sheet_name]
        recipe_data = {
            'name': sheet_name,
            'ingredients': [],
            'batch_weight': None,
            'loaf_weight': None,
            'formulas': []
        }

        # Scan the sheet for ingredient data
        # This is a simplified version - in reality, we'd need to understand
        # the specific layout of each sheet
        for row in ws.iter_rows(max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Look for common patterns
                    if 'Batch Weight' in str(cell.value):
                        # Try to get the value from adjacent cells
                        try:
                            batch_val = ws.cell(cell.row, cell.column + 1).value
                            if isinstance(batch_val, (int, float)):
                                recipe_data['batch_weight'] = batch_val
                        except:
                            pass

                    if 'grams' in str(cell.value).lower() and cell.column == 1:
                        # Might be a loaf weight specification
                        match = re.search(r'(\d+)\s*grams?', str(cell.value))
                        if match:
                            recipe_data['loaf_weight'] = int(match.group(1))

                # Capture formulas
                if cell.data_type == 'f':
                    recipe_data['formulas'].append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

        return recipe_data

    def extract_weekly_orders(self) -> Dict[str, Dict[str, int]]:
        """
        Extract weekly bread orders from the orders file
        Returns dict mapping bread names to daily quantities
        """
        if self.orders_wb is None:
            self.load_workbooks()

        # Look for the "Weekly Bread Totals" sheet
        if 'Weekly Bread Totals' not in self.orders_wb.sheetnames:
            print("Warning: 'Weekly Bread Totals' sheet not found")
            return {}

        ws = self.orders_wb['Weekly Bread Totals']
        orders = {}

        # Scan for bread totals
        # Structure: Item | Mon | Tue | Wed | Thu | Fri | Sat | Sun | Totals
        for row in ws.iter_rows(min_row=2, max_row=100):
            item_cell = row[0]
            if item_cell.value and isinstance(item_cell.value, str):
                item_name = item_cell.value.strip()

                # Skip headers and empty rows
                if item_name in ['Item', 'Monday', ''] or item_name is None:
                    continue

                # Get daily quantities (assuming Mon-Sun in columns B-H)
                daily_quantities = {}
                days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

                for i, day in enumerate(days):
                    try:
                        val = row[i + 1].value  # +1 because column A is item name
                        if isinstance(val, (int, float)):
                            daily_quantities[day] = int(val)
                        else:
                            daily_quantities[day] = 0
                    except:
                        daily_quantities[day] = 0

                if any(daily_quantities.values()):  # Only add if there are non-zero values
                    orders[item_name] = daily_quantities

        return orders

    def parse_formula_reference(self, formula: str) -> List[Tuple[str, str]]:
        """
        Parse Excel formula to extract sheet and cell references
        Example: "='Rustic White'!D3" -> [('Rustic White', 'D3')]
        """
        pattern = r"'?([^'!]+)'?!([A-Z]+\d+)"
        matches = re.findall(pattern, formula)
        return matches

    def get_cell_value(self, sheet_name: str, cell_ref: str, use_formula: bool = False):
        """Get value from a specific cell"""
        if self.formulas_wb is None:
            self.load_workbooks()

        if sheet_name not in self.formulas_wb.sheetnames:
            return None

        ws = self.formulas_wb[sheet_name]
        cell = ws[cell_ref]

        if use_formula and cell.data_type == 'f':
            return cell.value  # Return the formula
        else:
            # For data_only mode, we'd need to reload the workbook
            # For now, return the value
            return cell.value

    def close(self):
        """Close workbooks"""
        if self.formulas_wb:
            self.formulas_wb.close()
        if self.orders_wb:
            self.orders_wb.close()


def test_parser():
    """Test the parser"""
    parser = ExcelParser('Bread Formulas 2024.xlsx', 'Weekly Bread-Pastry Orders.xlsx')
    parser.load_workbooks()

    print("\n=== Recipe Sheets ===")
    recipes = parser.get_recipe_sheets()
    print(f"Found {len(recipes)} recipe sheets")
    for recipe in recipes[:5]:
        print(f"  - {recipe}")

    print("\n=== Starter Sheets ===")
    starters = parser.get_starter_sheets()
    for starter in starters:
        print(f"  - {starter}")

    print("\n=== Soaker Sheets ===")
    soakers = parser.get_soaker_sheets()
    for soaker in soakers:
        print(f"  - {soaker}")

    print("\n=== Sample Recipe Data ===")
    if recipes:
        recipe_data = parser.extract_recipe_data(recipes[0])
        print(f"Recipe: {recipe_data['name']}")
        print(f"Batch Weight: {recipe_data['batch_weight']}")
        print(f"Loaf Weight: {recipe_data['loaf_weight']}")
        print(f"Found {len(recipe_data['formulas'])} formulas")

    print("\n=== Weekly Orders ===")
    orders = parser.extract_weekly_orders()
    print(f"Found orders for {len(orders)} items")
    for item, quantities in list(orders.items())[:3]:
        total = sum(quantities.values())
        print(f"  {item}: {total} total")

    parser.close()


if __name__ == '__main__':
    test_parser()
