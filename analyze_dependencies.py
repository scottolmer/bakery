"""
Analyze which starters and soakers each bread recipe needs
"""
import openpyxl

wb = openpyxl.load_workbook('Bread Formulas 2024.xlsx', data_only=False)

# Analyze a few key bread recipes
recipes_to_analyze = ['Italian', 'Multigrain', 'Rustic White', 'Baguette', 'Pain dMie']

for recipe_name in recipes_to_analyze:
    if recipe_name not in wb.sheetnames:
        continue

    print(f"\n{'='*70}")
    print(f"ANALYZING: {recipe_name}")
    print(f"{'='*70}")

    ws = wb[recipe_name]

    # Look for ingredient references in the Final Dough section
    print("\nIngredients in Final Dough:")

    # Scan the sheet for ingredient names and amounts
    for row_idx in range(1, 50):
        for col_idx in range(1, 6):  # Check first 5 columns
            cell = ws.cell(row_idx, col_idx)

            if cell.value:
                cell_str = str(cell.value).lower()

                # Look for starter keywords
                starter_keywords = ['levain', 'poolish', 'biga', 'starter']
                soaker_keywords = ['soaker', 'soak']

                for keyword in starter_keywords:
                    if keyword in cell_str:
                        # Get the amount from next column
                        amount_cell = ws.cell(row_idx, col_idx + 1)
                        if amount_cell.value:
                            # Check if it's a formula or value
                            if amount_cell.data_type == 'f':
                                print(f"  STARTER: {cell.value}")
                                print(f"    Location: {cell.coordinate}")
                                print(f"    Formula: {amount_cell.value}")
                            elif isinstance(amount_cell.value, (int, float)):
                                print(f"  STARTER: {cell.value} = {amount_cell.value}g")
                                print(f"    Location: {cell.coordinate}")
                        break

                for keyword in soaker_keywords:
                    if keyword in cell_str:
                        amount_cell = ws.cell(row_idx, col_idx + 1)
                        if amount_cell.value:
                            if amount_cell.data_type == 'f':
                                print(f"  SOAKER: {cell.value}")
                                print(f"    Location: {cell.coordinate}")
                                print(f"    Formula: {amount_cell.value}")
                            elif isinstance(amount_cell.value, (int, float)):
                                print(f"  SOAKER: {cell.value} = {amount_cell.value}g")
                                print(f"    Location: {cell.coordinate}")
                        break

    # Look for batch weight
    print("\nBatch Information:")
    for row_idx in range(1, 50):
        for col_idx in range(1, 6):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and 'batch weight' in str(cell.value).lower():
                amount_cell = ws.cell(row_idx, col_idx + 1)
                if amount_cell.value:
                    print(f"  Batch Weight: {amount_cell.value}g at {amount_cell.coordinate}")

wb.close()

print("\n" + "="*70)
print("ANALYSIS COMPLETE")
print("="*70)
