import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('Weekly Bread-Pastry Orders.xlsx', data_only=False)
print('Available sheets:', wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    print()

    for row_idx, row in enumerate(ws.iter_rows(max_row=30), 1):
        row_data = []
        for cell in row:
            if cell.value is None:
                continue
            if cell.data_type == 'f':
                row_data.append(f'{cell.coordinate}: {cell.value} (Formula)')
            else:
                row_data.append(f'{cell.coordinate}: {cell.value}')
        if row_data:
            print(f'Row {row_idx}: {", ".join(row_data)}')
