"""
Analyze the bakery workflow sheets
"""
import openpyxl

wb = openpyxl.load_workbook('Bread Formulas 2024.xlsx', data_only=True)

print("="*70)
print("EXPERIMENTAL MIXING SHEET (Today's Mix)")
print("="*70)
ws = wb['Experimental Mixing Sheet']
print(f"Dimensions: {ws.dimensions}")
print("\nFirst 30 rows with content:")
for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 9):  # First 8 columns
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            row_data.append(f"{chr(64+col_idx)}{row_idx}:{cell.value}")
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n" + "="*70)
print("EXPERIMENTAL STARTERS SHEET (Prep for Tomorrow)")
print("="*70)
ws_starters = wb['Experimental Starters']
print(f"Dimensions: {ws_starters.dimensions}")
print("\nFirst 30 rows with content:")
for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 9):
        cell = ws_starters.cell(row_idx, col_idx)
        if cell.value:
            row_data.append(f"{chr(64+col_idx)}{row_idx}:{cell.value}")
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n" + "="*70)
print("EXPERIMENTAL SOAKS SHEET (Prep for Tomorrow)")
print("="*70)
ws_soaks = wb['Experimental Soaks']
print(f"Dimensions: {ws_soaks.dimensions}")
print("\nFirst 30 rows with content:")
for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 9):
        cell = ws_soaks.cell(row_idx, col_idx)
        if cell.value:
            row_data.append(f"{chr(64+col_idx)}{row_idx}:{cell.value}")
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n" + "="*70)
print("MEP HUB (Control Panel)")
print("="*70)
ws_hub = wb['MEP Hub']
print(f"Dimensions: {ws_hub.dimensions}")
print("\nFirst 30 rows with content:")
for row_idx in range(1, 31):
    row_data = []
    for col_idx in range(1, 9):
        cell = ws_hub.cell(row_idx, col_idx)
        if cell.value:
            row_data.append(f"{chr(64+col_idx)}{row_idx}:{cell.value}")
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data)}")

wb.close()
