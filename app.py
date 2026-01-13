"""
Bakery Production Management System
Main Flask application
"""
import os
import json
from flask import Flask, render_template, jsonify, request
from flask_cors import CORS
from config import Config
from models import db, Recipe, Ingredient, RecipeIngredient, ProductionRun, ProductionItem, ProductionIngredient, ScheduleTemplate, MixerCapacity, Customer, Order, WeeklyOrderTemplate, MixingLog, MixingLogEntry, DDTTarget, ProductionIssue, InventoryTransaction
from datetime import datetime, date, timedelta
from mep_calculator import MEPCalculator

app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
CORS(app)
db.init_app(app)


# =============================================================================
# Database Initialization
# =============================================================================

@app.cli.command('init-db')
def init_db():
    """Initialize the database"""
    db.create_all()
    print("Database tables created successfully!")


@app.cli.command('seed-db')
def seed_db():
    """Seed database with sample data"""
    print("Seeding database...")

    # Create base ingredients
    ingredients_data = [
        ('Red Rose Flour', 'flour'),
        ('Whole Wheat Flour', 'flour'),
        ('Water', 'water'),
        ('Salt', 'salt'),
        ('Levain', 'starter'),
        ('Poolish', 'starter'),
        ('Biga', 'starter'),
        ('Olive Oil', 'oil'),
        ('Yeast', 'yeast'),
        ('7 Grain Soaker', 'soaker'),
        ('RW Soaker', 'soaker'),
    ]

    for name, category in ingredients_data:
        if not Ingredient.query.filter_by(name=name).first():
            ingredient = Ingredient(name=name, category=category)
            db.session.add(ingredient)

    db.session.commit()
    print(f"Created {len(ingredients_data)} ingredients")

    # Create sample recipes
    # Italian Bread
    italian = Recipe(
        name='Italian',
        recipe_type='bread',
        base_batch_weight=1000,
        loaf_weight=1000
    )
    db.session.add(italian)
    db.session.commit()

    # Add ingredients to Italian recipe
    red_rose = Ingredient.query.filter_by(name='Red Rose Flour').first()
    water = Ingredient.query.filter_by(name='Water').first()
    salt = Ingredient.query.filter_by(name='Salt').first()
    levain = Ingredient.query.filter_by(name='Levain').first()

    recipe_ingredients = [
        RecipeIngredient(recipe_id=italian.id, ingredient_id=red_rose.id, percentage=70.0, is_percentage=True, order=1),
        RecipeIngredient(recipe_id=italian.id, ingredient_id=water.id, percentage=65.0, is_percentage=True, order=2),
        RecipeIngredient(recipe_id=italian.id, ingredient_id=salt.id, percentage=2.0, is_percentage=True, order=3),
        RecipeIngredient(recipe_id=italian.id, ingredient_id=levain.id, percentage=30.0, is_percentage=True, order=4),
    ]

    for ri in recipe_ingredients:
        db.session.add(ri)

    db.session.commit()
    print("Created sample Italian bread recipe")

    # Create mixer capacity limit
    capacity = MixerCapacity(recipe_id=italian.id, max_batch_weight=5000)
    db.session.add(capacity)
    db.session.commit()

    print("Database seeded successfully!")


@app.cli.command('import-recipes')
def import_recipes():
    """Import recipes from Excel file"""
    import openpyxl
    from config import Config

    print("Importing recipes from Excel...")

    try:
        wb = openpyxl.load_workbook(Config.BREAD_FORMULAS_FILE, data_only=True)
    except FileNotFoundError:
        print(f"Error: Could not find {Config.BREAD_FORMULAS_FILE}")
        return

    # Define which sheets to import and their types
    bread_sheets = ['Italian', 'Multigrain', 'Rustic White', 'Baguette', 'Pain dMie',
                    'Miche', 'Brioche', 'Schiacciata', 'Stollen', 'Pumpkin Miche',
                    'Dinkel', 'Hot Cross Buns', 'Fino', 'Focaccia', 'Croissant',
                    'Brotchen', 'Chocolate Croissant', 'Light Rye ', 'Dark Rye',
                    'Ciabatta', 'Naan', 'Irish Soda Bread ']

    starter_sheets = ['Levain', 'Itl Levain', 'Biga', 'Emmy(starter)', 'Poolish']
    soaker_sheets = ['RW Soaker', '7 Grain Soaker', 'Dinkel Soaker']

    recipes_imported = 0
    ingredients_created = 0

    def import_sheet(sheet_name, recipe_type):
        nonlocal recipes_imported, ingredients_created

        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} - sheet not found")
            return

        sheet = wb[sheet_name]

        # Extract recipe name from row 1, column C
        recipe_name = None
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[2]:  # Column C
                recipe_name = str(row[2]).strip()
                if recipe_name and recipe_name not in ['Overall', 'Ingredient', 'Bakers %']:
                    break

        if not recipe_name:
            print(f"  Skipping {sheet_name} - could not find recipe name")
            return

        # Check if recipe already exists
        existing = Recipe.query.filter_by(name=recipe_name).first()
        if existing:
            print(f"  Skipping {recipe_name} - already exists")
            return

        # Extract loaf weight (look for gram/g in first few rows)
        loaf_weight = 1000  # default
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[2] and isinstance(row[2], str):
                text = str(row[2]).lower()
                if 'gram' in text or 'g ' in text:
                    # Try to extract number
                    import re
                    match = re.search(r'(\d+)', text)
                    if match:
                        loaf_weight = int(match.group(1))
                        break
            # Also check column D for numeric values
            if row[3] and isinstance(row[3], (int, float)):
                if 500 <= row[3] <= 5000:  # reasonable batch size
                    loaf_weight = int(row[3])
                    break

        # Find ingredients section (look for "Bakers %" header)
        ingredients = []
        start_row = None

        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row[1] and str(row[1]).strip() in ['Bakers %', "Bakers'%"]:
                start_row = i + 1  # Ingredients start on next row
                break

        if not start_row:
            print(f"  Skipping {recipe_name} - could not find ingredients section")
            return

        # Extract ingredients
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            baker_pct = row[1]  # Column B
            ing_name = row[2]   # Column C

            # Stop at sum row or empty rows
            if not ing_name or not baker_pct:
                continue

            if isinstance(baker_pct, str) and 'SUM' in baker_pct.upper():
                break

            # Clean ingredient name
            ing_name = str(ing_name).strip()
            if ing_name in ['', 'Ingredient', 'Overall', 'Poolish', 'Biga', 'Levain Build', 'Soaker']:
                continue

            # Convert percentage
            try:
                pct = float(baker_pct) * 100  # Convert 0.72 to 72%
                if pct > 0 and pct < 500:  # Sanity check
                    ingredients.append((ing_name, pct))
            except (ValueError, TypeError):
                continue

        if not ingredients:
            print(f"  Skipping {recipe_name} - no ingredients found")
            return

        # Create recipe
        recipe = Recipe(
            name=recipe_name,
            recipe_type=recipe_type,
            base_batch_weight=loaf_weight,
            loaf_weight=loaf_weight
        )
        db.session.add(recipe)
        db.session.flush()

        # Create/find ingredients and link to recipe
        for order, (ing_name, pct) in enumerate(ingredients, 1):
            # Find or create ingredient
            ingredient = Ingredient.query.filter_by(name=ing_name).first()
            if not ingredient:
                # Determine category
                category = 'other'
                name_lower = ing_name.lower()
                if 'flour' in name_lower or 'rose' in name_lower or 'wheat' in name_lower:
                    category = 'flour'
                elif 'water' in name_lower:
                    category = 'water'
                elif 'salt' in name_lower:
                    category = 'salt'
                elif 'yeast' in name_lower:
                    category = 'yeast'
                elif 'oil' in name_lower:
                    category = 'oil'
                elif 'levain' in name_lower or 'poolish' in name_lower or 'biga' in name_lower or 'emmy' in name_lower or 'starter' in name_lower:
                    category = 'starter'
                elif 'soaker' in name_lower or 'grain' in name_lower:
                    category = 'soaker'

                ingredient = Ingredient(name=ing_name, category=category)
                db.session.add(ingredient)
                db.session.flush()
                ingredients_created += 1

            # Link ingredient to recipe
            recipe_ingredient = RecipeIngredient(
                recipe_id=recipe.id,
                ingredient_id=ingredient.id,
                percentage=pct,
                is_percentage=True,
                order=order
            )
            db.session.add(recipe_ingredient)

        db.session.commit()
        recipes_imported += 1
        print(f"  ✓ Imported {recipe_name} ({len(ingredients)} ingredients)")

    # Import breads
    print("\nImporting bread recipes...")
    for sheet_name in bread_sheets:
        import_sheet(sheet_name, 'bread')

    # Import starters
    print("\nImporting starter recipes...")
    for sheet_name in starter_sheets:
        import_sheet(sheet_name, 'starter')

    # Import soakers
    print("\nImporting soaker recipes...")
    for sheet_name in soaker_sheets:
        import_sheet(sheet_name, 'soaker')

    print(f"\n✅ Import complete!")
    print(f"   Recipes imported: {recipes_imported}")
    print(f"   New ingredients created: {ingredients_created}")


# =============================================================================
# API Routes
# =============================================================================

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')


@app.route('/history')
def history():
    """Production history page"""
    return render_template('history.html')


@app.route('/production')
def production():
    """Production entry page"""
    return render_template('production.html')


@app.route('/recipes')
def recipes_page():
    """Recipe management page"""
    return render_template('recipes.html')


@app.route('/mep')
def mep_complete():
    """Complete MEP sheets page"""
    return render_template('mep_complete.html')


@app.route('/orders')
def orders_page():
    """Orders management page"""
    return render_template('orders.html')


@app.route('/orders/weekly')
def weekly_orders_page():
    """Weekly order sheets page"""
    return render_template('weekly_orders.html')


@app.route('/api/recipes', methods=['GET'])
def get_recipes():
    """Get all recipes"""
    recipes = Recipe.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': r.id,
        'name': r.name,
        'recipe_type': r.recipe_type,
        'base_batch_weight': r.base_batch_weight,
        'loaf_weight': r.loaf_weight,
        'selling_price': r.selling_price
    } for r in recipes])


@app.route('/api/recipes/<int:recipe_id>', methods=['GET'])
def get_recipe(recipe_id):
    """Get a specific recipe with ingredients"""
    recipe = Recipe.query.get_or_404(recipe_id)
    ingredients = []

    for ri in recipe.ingredients:
        ingredients.append({
            'ingredient_id': ri.ingredient_id,
            'ingredient_name': ri.ingredient.name,
            'percentage': ri.percentage,
            'amount_grams': ri.amount_grams,
            'is_percentage': ri.is_percentage,
            'category': ri.ingredient.category
        })

    # Calculate recipe cost based on ingredients
    recipe_cost_per_loaf = 0
    for ri in recipe.ingredients:
        if ri.ingredient.cost_per_unit:
            if ri.is_percentage:
                # Calculate actual grams from percentage
                # Percentage is based on flour weight (100% = base_batch_weight)
                actual_grams = (ri.percentage / 100) * recipe.base_batch_weight
            else:
                actual_grams = ri.amount_grams

            # Cost for this ingredient in the batch
            ingredient_cost = actual_grams * ri.ingredient.cost_per_unit

            # Divide by number of loaves in batch
            loaves_per_batch = recipe.base_batch_weight / recipe.loaf_weight
            recipe_cost_per_loaf += ingredient_cost / loaves_per_batch

    return jsonify({
        'id': recipe.id,
        'name': recipe.name,
        'recipe_type': recipe.recipe_type,
        'base_batch_weight': recipe.base_batch_weight,
        'loaf_weight': recipe.loaf_weight,
        'selling_price': recipe.selling_price,
        'cost_per_loaf': round(recipe_cost_per_loaf, 2),
        'ingredients': ingredients,
        'notes': recipe.notes
    })


@app.route('/api/production/calculate', methods=['POST'])
def calculate_production():
    """
    Calculate ingredient amounts for production
    Expected input: { "items": [{"recipe_id": 1, "quantity": 10}, ...] }
    """
    data = request.json
    items = data.get('items', [])

    results = []

    for item in items:
        recipe_id = item['recipe_id']
        quantity = item['quantity']

        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            continue

        # Calculate batch weight needed
        total_weight = quantity * recipe.loaf_weight

        # Calculate ingredients
        ingredients = []
        for ri in recipe.ingredients:
            if ri.is_percentage:
                # Baker's percentage calculation
                amount = (ri.percentage / 100.0) * recipe.base_batch_weight * (total_weight / recipe.base_batch_weight)
            else:
                # Fixed amount
                amount = ri.amount_grams * (total_weight / recipe.base_batch_weight)

            ingredients.append({
                'name': ri.ingredient.name,
                'amount_grams': round(amount, 1),
                'category': ri.ingredient.category
            })

        results.append({
            'recipe_name': recipe.name,
            'quantity': quantity,
            'total_weight': total_weight,
            'ingredients': ingredients
        })

    return jsonify(results)


@app.route('/api/production/history', methods=['GET'])
def get_production_history():
    """Get production history with optional date filtering"""
    # Query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    recipe_name = request.args.get('recipe_name')

    query = ProductionRun.query

    if start_date:
        query = query.filter(ProductionRun.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(ProductionRun.date <= datetime.strptime(end_date, '%Y-%m-%d').date())

    production_runs = query.order_by(ProductionRun.date.desc()).all()

    results = []
    for run in production_runs:
        items = []
        for item in run.items:
            if recipe_name and item.recipe.name != recipe_name:
                continue

            items.append({
                'recipe_name': item.recipe.name,
                'quantity': item.quantity,
                'batch_weight': item.batch_weight
            })

        if items or not recipe_name:  # Include run if it has matching items or no filter
            results.append({
                'id': run.id,
                'date': run.date.isoformat(),
                'batch_id': run.batch_id,
                'created_by': run.created_by,
                'items': items
            })

    return jsonify(results)


@app.route('/api/production/batch/<batch_id>', methods=['GET'])
def get_production_by_batch(batch_id):
    """Get production run by batch ID"""
    production_run = ProductionRun.query.filter_by(batch_id=batch_id).first()

    if not production_run:
        return jsonify({'error': 'No production run found for this batch ID'}), 404

    items = []
    for item in production_run.items:
        items.append({
            'recipe_name': item.recipe.name,
            'quantity': item.quantity,
            'batch_weight': item.batch_weight
        })

    result = {
        'id': production_run.id,
        'date': production_run.date.isoformat(),
        'batch_id': production_run.batch_id,
        'created_by': production_run.created_by,
        'notes': production_run.notes,
        'items': items
    }

    return jsonify(result)


@app.route('/api/production/save', methods=['POST'])
def save_production():
    """Save a production run"""
    data = request.json

    run_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    production_run = ProductionRun(
        date=run_date,
        batch_id=run_date.strftime('%m%d%y'),  # Auto-generate batch ID (MMDDYY)
        created_by=data.get('created_by', 'admin'),
        notes=data.get('notes', '')
    )
    db.session.add(production_run)
    db.session.flush()  # Get the ID

    for item_data in data['items']:
        recipe = Recipe.query.get(item_data['recipe_id'])

        production_item = ProductionItem(
            production_run_id=production_run.id,
            recipe_id=item_data['recipe_id'],
            customer_id=item_data.get('customer_id'),
            quantity=item_data['quantity'],
            batch_weight=item_data['quantity'] * recipe.loaf_weight
        )
        db.session.add(production_item)

    db.session.commit()

    return jsonify({
        'success': True,
        'production_run_id': production_run.id,
        'batch_id': production_run.batch_id
    })


@app.route('/api/mep/<date_str>')
def get_mep_sheet(date_str):
    """Generate MEP (mise en place) sheet for a specific date"""
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # Find production run for this date
    production_run = ProductionRun.query.filter_by(date=target_date).first()

    if not production_run:
        return jsonify({'error': 'No production run found for this date'}), 404

    # Generate MEP sheet data
    mep_data = {
        'date': target_date.isoformat(),
        'breads': [],
        'total_ingredients': {}
    }

    for item in production_run.items:
        recipe = item.recipe
        bread_data = {
            'name': recipe.name,
            'quantity': item.quantity,
            'loaf_weight': recipe.loaf_weight,
            'total_weight': item.batch_weight,
            'ingredients': []
        }

        # Calculate ingredients for this bread
        for ri in recipe.ingredients:
            if ri.is_percentage:
                amount = (ri.percentage / 100.0) * item.batch_weight
            else:
                amount = ri.amount_grams * (item.batch_weight / recipe.base_batch_weight)

            bread_data['ingredients'].append({
                'name': ri.ingredient.name,
                'amount': round(amount, 1),
                'category': ri.ingredient.category
            })

            # Add to total ingredients
            ingredient_name = ri.ingredient.name
            if ingredient_name not in mep_data['total_ingredients']:
                mep_data['total_ingredients'][ingredient_name] = 0
            mep_data['total_ingredients'][ingredient_name] += amount

        mep_data['breads'].append(bread_data)

    # Round total ingredients
    for ingredient in mep_data['total_ingredients']:
        mep_data['total_ingredients'][ingredient] = round(mep_data['total_ingredients'][ingredient], 1)

    return jsonify(mep_data)


@app.route('/api/ingredients', methods=['GET'])
def get_ingredients():
    """Get all ingredients"""
    ingredients = Ingredient.query.all()
    return jsonify([{
        'id': i.id,
        'name': i.name,
        'category': i.category,
        'unit': i.unit,
        'cost_per_unit': i.cost_per_unit
    } for i in ingredients])


@app.route('/api/ingredients', methods=['POST'])
def create_ingredient():
    """Create a new ingredient"""
    data = request.json

    # Check if ingredient already exists
    existing = Ingredient.query.filter_by(name=data['name']).first()
    if existing:
        return jsonify({'error': 'Ingredient already exists'}), 400

    ingredient = Ingredient(
        name=data['name'],
        category=data.get('category', 'other'),
        unit=data.get('unit', 'grams'),
        cost_per_unit=data.get('cost_per_unit')
    )
    db.session.add(ingredient)
    db.session.commit()

    return jsonify({
        'success': True,
        'ingredient_id': ingredient.id
    })


@app.route('/api/ingredients/<int:ingredient_id>', methods=['DELETE'])
def delete_ingredient(ingredient_id):
    """Delete an ingredient"""
    ingredient = Ingredient.query.get_or_404(ingredient_id)

    # Check if ingredient is used in any recipes
    recipes_using = RecipeIngredient.query.filter_by(ingredient_id=ingredient_id).count()
    if recipes_using > 0:
        return jsonify({
            'error': f'Cannot delete ingredient. It is used in {recipes_using} recipe(s).'
        }), 400

    # Delete any inventory transactions
    InventoryTransaction.query.filter_by(ingredient_id=ingredient_id).delete()

    # Delete the ingredient
    db.session.delete(ingredient)
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/recipes', methods=['POST'])
def create_recipe():
    """Create a new recipe"""
    data = request.json

    # Check if recipe already exists
    existing = Recipe.query.filter_by(name=data['name']).first()
    if existing:
        return jsonify({'error': 'Recipe already exists'}), 400

    recipe = Recipe(
        name=data['name'],
        recipe_type=data.get('recipe_type', 'bread'),
        base_batch_weight=data.get('base_batch_weight'),
        loaf_weight=data.get('loaf_weight'),
        notes=data.get('notes', ''),
        selling_price=data.get('selling_price')
    )
    db.session.add(recipe)
    db.session.flush()  # Get recipe.id before adding ingredients

    # Add ingredients if provided
    if 'ingredients' in data and data['ingredients']:
        for ing_data in data['ingredients']:
            recipe_ingredient = RecipeIngredient(
                recipe_id=recipe.id,
                ingredient_id=ing_data['ingredient_id'],
                percentage=ing_data.get('percentage'),
                is_percentage=ing_data.get('is_percentage', True),
                amount_grams=ing_data.get('amount_grams')
            )
            db.session.add(recipe_ingredient)

    db.session.commit()

    return jsonify({
        'success': True,
        'recipe_id': recipe.id
    })


@app.route('/api/recipes/<int:recipe_id>', methods=['PUT'])
def update_recipe(recipe_id):
    """Update a recipe"""
    recipe = Recipe.query.get_or_404(recipe_id)
    data = request.json

    recipe.name = data.get('name', recipe.name)
    recipe.recipe_type = data.get('recipe_type', recipe.recipe_type)
    recipe.base_batch_weight = data.get('base_batch_weight', recipe.base_batch_weight)
    recipe.loaf_weight = data.get('loaf_weight', recipe.loaf_weight)
    recipe.notes = data.get('notes', recipe.notes)
    recipe.selling_price = data.get('selling_price', recipe.selling_price)

    # Update ingredients if provided
    if 'ingredients' in data:
        # Delete existing ingredients
        RecipeIngredient.query.filter_by(recipe_id=recipe_id).delete()

        # Add new ingredients
        for ing_data in data['ingredients']:
            recipe_ingredient = RecipeIngredient(
                recipe_id=recipe.id,
                ingredient_id=ing_data['ingredient_id'],
                percentage=ing_data.get('percentage'),
                is_percentage=ing_data.get('is_percentage', True),
                amount_grams=ing_data.get('amount_grams')
            )
            db.session.add(recipe_ingredient)

    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/recipes/<int:recipe_id>', methods=['DELETE'])
def delete_recipe(recipe_id):
    """Delete a recipe"""
    recipe = Recipe.query.get_or_404(recipe_id)
    recipe.is_active = False
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/recipes/<int:recipe_id>/ingredients', methods=['POST'])
def add_recipe_ingredient(recipe_id):
    """Add an ingredient to a recipe"""
    recipe = Recipe.query.get_or_404(recipe_id)
    data = request.json

    recipe_ingredient = RecipeIngredient(
        recipe_id=recipe_id,
        ingredient_id=data['ingredient_id'],
        percentage=data.get('percentage'),
        amount_grams=data.get('amount_grams'),
        is_percentage=data.get('is_percentage', True),
        order=data.get('order', 0)
    )
    db.session.add(recipe_ingredient)
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/mep/calculate', methods=['POST'])
def calculate_mep_sheets():
    """
    Calculate all four MEP sheets for given production items
    Input: { "items": [{"recipe_id": 1, "quantity": 10}, ...] }
    Output: {
        "mix_sheet": {...},
        "starter_sheet": {...},
        "soak_sheet": {...},
        "mep_ingredients": {...}
    }
    """
    data = request.json
    items = data.get('items', [])

    if not items:
        return jsonify({'error': 'No items provided'}), 400

    calculator = MEPCalculator(items)
    sheets = calculator.calculate_all_sheets()

    return jsonify(sheets)


@app.route('/api/mep/<date_str>/all', methods=['GET'])
def get_all_mep_sheets(date_str):
    """
    Get all MEP sheets for a specific delivery date
    System calculates backwards: delivery date -> mix date -> prep date
    """
    try:
        delivery_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # The date entered is the DELIVERY date
    # Calculate backwards:
    # - Mix date: delivery - 1 day (when doughs are mixed)
    # - Prep date: delivery - 2 days (when starters/soakers are built)
    mix_date = delivery_date - timedelta(days=1)
    prep_date = delivery_date - timedelta(days=2)

    # Find production run for the delivery date
    production_run = ProductionRun.query.filter_by(date=delivery_date).first()

    if not production_run:
        return jsonify({'error': 'No production run found for this delivery date'}), 404

    # Build items list
    items = []
    for item in production_run.items:
        items.append({
            'recipe_id': item.recipe_id,
            'quantity': item.quantity
        })

    # Calculate all sheets (pass delivery_date for Emmy Feed calculation)
    calculator = MEPCalculator(items, delivery_date=delivery_date)
    sheets = calculator.calculate_all_sheets()

    # Add metadata with proper timeline
    # All calculated from the delivery date
    sheets['delivery_date'] = delivery_date.isoformat()
    sheets['mix_date'] = mix_date.isoformat()
    sheets['prep_date'] = prep_date.isoformat()
    sheets['batch_id'] = production_run.batch_id  # Add batch ID for labeling

    return jsonify(sheets)


@app.route('/api/customers', methods=['GET'])
def get_customers():
    """Get all active customers"""
    customers = Customer.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'short_name': c.short_name
    } for c in customers])


@app.route('/api/customers', methods=['POST'])
def create_customer():
    """Create a new customer"""
    data = request.json

    # Check if customer already exists
    existing = Customer.query.filter_by(name=data['name']).first()
    if existing:
        return jsonify({'error': 'Customer already exists'}), 400

    customer = Customer(
        name=data['name'],
        short_name=data.get('short_name', data['name'][:10])
    )
    db.session.add(customer)
    db.session.commit()

    return jsonify({
        'success': True,
        'customer_id': customer.id
    })


@app.route('/api/customers/<int:customer_id>', methods=['DELETE'])
def delete_customer(customer_id):
    """Delete a customer"""
    customer = Customer.query.get_or_404(customer_id)
    customer.is_active = False
    db.session.commit()

    return jsonify({'success': True})


@app.route('/customers')
def customers_page():
    """Customer management page"""
    return render_template('customers.html')


@app.route('/orders/quick')
def quick_order_page():
    """Quick multi-day order entry page"""
    return render_template('quick_order.html')


@app.route('/api/customers/<int:customer_id>/typical-breads', methods=['GET'])
def get_customer_typical_breads(customer_id):
    """Get breads this customer typically orders based on order history"""
    # Get unique recipes from customer's order history
    orders = db.session.query(
        Order.recipe_id,
        Recipe.name.label('recipe_name'),
        db.func.count(Order.id).label('order_count')
    ).join(Recipe).filter(
        Order.customer_id == customer_id
    ).group_by(Order.recipe_id, Recipe.name).order_by(
        db.func.count(Order.id).desc()
    ).all()

    breads = [{
        'recipe_id': order.recipe_id,
        'recipe_name': order.recipe_name,
        'order_count': order.order_count
    } for order in orders]

    return jsonify({'breads': breads})


@app.route('/api/orders/bulk', methods=['POST'])
def create_bulk_orders():
    """Create multiple orders at once"""
    data = request.json
    orders_data = data.get('orders', [])

    if not orders_data:
        return jsonify({'error': 'No orders provided'}), 400

    created_count = 0
    updated_count = 0

    for order_data in orders_data:
        # Check if order already exists for this customer/recipe/date
        existing_order = Order.query.filter_by(
            customer_id=order_data['customer_id'],
            recipe_id=order_data['recipe_id'],
            order_date=datetime.strptime(order_data['order_date'], '%Y-%m-%d').date()
        ).first()

        if existing_order:
            # Update existing order
            existing_order.quantity = order_data['quantity']
            updated_count += 1
        else:
            # Create new order
            order = Order(
                customer_id=order_data['customer_id'],
                recipe_id=order_data['recipe_id'],
                order_date=datetime.strptime(order_data['order_date'], '%Y-%m-%d').date(),
                quantity=order_data['quantity'],
                day_of_week=datetime.strptime(order_data['order_date'], '%Y-%m-%d').strftime('%A')
            )
            db.session.add(order)
            created_count += 1

    db.session.commit()

    return jsonify({
        'success': True,
        'created': created_count,
        'updated': updated_count,
        'total': created_count + updated_count
    })


@app.route('/api/orders', methods=['GET'])
def get_orders():
    """Get all orders with optional date and customer filtering"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    customer_id = request.args.get('customer_id')

    query = Order.query

    if start_date:
        query = query.filter(Order.order_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Order.order_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if customer_id:
        query = query.filter(Order.customer_id == int(customer_id))

    orders = query.order_by(Order.order_date.asc()).all()

    return jsonify([{
        'id': o.id,
        'customer_id': o.customer_id,
        'customer_name': o.customer.name,
        'customer_short_name': o.customer.short_name,
        'recipe_id': o.recipe_id,
        'recipe_name': o.recipe.name,
        'order_date': o.order_date.isoformat(),
        'day_of_week': o.day_of_week,
        'quantity': o.quantity,
        'notes': o.notes
    } for o in orders])


@app.route('/api/orders', methods=['POST'])
def create_order():
    """Create a new order"""
    data = request.json

    order = Order(
        customer_id=data['customer_id'],
        recipe_id=data['recipe_id'],
        order_date=datetime.strptime(data['order_date'], '%Y-%m-%d').date(),
        quantity=data['quantity'],
        day_of_week=data.get('day_of_week', ''),
        notes=data.get('notes', '')
    )

    db.session.add(order)
    db.session.commit()

    return jsonify({
        'success': True,
        'order_id': order.id
    })


@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    """Delete an order"""
    order = Order.query.get_or_404(order_id)
    db.session.delete(order)
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/orders/delete-week', methods=['POST'])
def delete_week_orders():
    """Delete all orders for a customer in a date range"""
    data = request.json
    customer_id = data['customer_id']
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

    # Delete orders in this range for this customer
    orders = Order.query.filter(
        Order.customer_id == customer_id,
        Order.order_date >= start_date,
        Order.order_date <= end_date
    ).all()

    for order in orders:
        db.session.delete(order)

    db.session.commit()

    return jsonify({'success': True, 'deleted': len(orders)})


@app.route('/api/orders/aggregate', methods=['POST'])
def aggregate_orders():
    """
    Aggregate orders by date and recipe to calculate production needs
    Input: { "start_date": "2026-01-06", "end_date": "2026-01-12" }
    Output: { "2026-01-06": [{"recipe_id": 1, "recipe_name": "Italian", "total_quantity": 50}, ...], ... }
    """
    data = request.json
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

    # Get all orders in date range
    orders = Order.query.filter(
        Order.order_date >= start_date,
        Order.order_date <= end_date
    ).all()

    # Aggregate by date and recipe
    aggregated = {}
    for order in orders:
        date_key = order.order_date.isoformat()

        if date_key not in aggregated:
            aggregated[date_key] = {}

        if order.recipe_id not in aggregated[date_key]:
            aggregated[date_key][order.recipe_id] = {
                'recipe_id': order.recipe_id,
                'recipe_name': order.recipe.name,
                'total_quantity': 0
            }

        aggregated[date_key][order.recipe_id]['total_quantity'] += order.quantity

    # Convert to list format
    result = {}
    for date_key, recipes in aggregated.items():
        result[date_key] = list(recipes.values())

    return jsonify(result)


@app.route('/api/orders/create-production', methods=['POST'])
def create_production_from_orders():
    """
    Auto-create production runs from aggregated orders
    Input: { "start_date": "2026-01-06", "end_date": "2026-01-12" }
    Output: { "success": true, "production_runs_created": 5 }
    """
    data = request.json
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

    # Get aggregated orders
    orders = Order.query.filter(
        Order.order_date >= start_date,
        Order.order_date <= end_date
    ).all()

    # Aggregate by date and recipe
    aggregated = {}
    for order in orders:
        date_key = order.order_date

        if date_key not in aggregated:
            aggregated[date_key] = {}

        if order.recipe_id not in aggregated[date_key]:
            aggregated[date_key][order.recipe_id] = 0

        aggregated[date_key][order.recipe_id] += order.quantity

    # Create production runs
    runs_created = 0
    for production_date, recipes in aggregated.items():
        # Check if production run already exists
        existing_run = ProductionRun.query.filter_by(date=production_date).first()

        if existing_run:
            # Delete existing items and recreate
            for item in existing_run.items:
                db.session.delete(item)
            production_run = existing_run
        else:
            # Create new production run
            production_run = ProductionRun(
                date=production_date,
                batch_id=production_date.strftime('%m%d%y'),  # Auto-generate batch ID (MMDDYY)
                created_by='orders_system',
                notes='Auto-generated from customer orders'
            )
            db.session.add(production_run)
            db.session.flush()
            runs_created += 1

        # Add production items
        for recipe_id, quantity in recipes.items():
            recipe = Recipe.query.get(recipe_id)
            production_item = ProductionItem(
                production_run_id=production_run.id,
                recipe_id=recipe_id,
                quantity=quantity,
                batch_weight=quantity * recipe.loaf_weight
            )
            db.session.add(production_item)

    db.session.commit()

    return jsonify({
        'success': True,
        'production_runs_created': runs_created
    })


# =============================================================================
# Mixing Log API Endpoints
# =============================================================================

@app.route('/api/ddt-targets')
def get_ddt_targets():
    """Get all DDT target ranges for validation"""
    targets = DDTTarget.query.filter_by(is_active=True).all()

    return jsonify({
        'targets': [{
            'id': target.id,
            'bread_name': target.bread_name,
            'target_temp_min': target.target_temp_min,
            'target_temp_max': target.target_temp_max,
            'notes': target.notes
        } for target in targets]
    })


@app.route('/api/mixing-log/breads/<date_str>')
def get_mixing_log_breads(date_str):
    """Get list of breads from production run for auto-population"""
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # Calculate mix date (delivery date - 1 day)
    mix_date = target_date - timedelta(days=1)

    # Find production run for this mix date
    production_run = ProductionRun.query.filter_by(date=target_date).first()
    if not production_run:
        return jsonify({'error': f'No production run found for date {date_str}'}), 404

    # Get breads from production run
    breads = []
    for item in production_run.items:
        recipe = item.recipe

        # Get DDT target for this bread
        ddt_target = DDTTarget.query.filter_by(bread_name=recipe.name, is_active=True).first()

        breads.append({
            'recipe_id': recipe.id,
            'bread_name': recipe.name,
            'quantity': item.quantity,
            'batch_weight': item.batch_weight,
            'ddt_target': {
                'min': ddt_target.target_temp_min if ddt_target else None,
                'max': ddt_target.target_temp_max if ddt_target else None
            } if ddt_target else None
        })

    return jsonify({
        'date': target_date.strftime('%Y-%m-%d'),
        'mix_date': mix_date.strftime('%Y-%m-%d'),
        'production_run_id': production_run.id,
        'batch_id': production_run.batch_id,
        'breads': breads
    })


@app.route('/api/mixing-log/<date_str>')
def get_mixing_log(date_str):
    """Get mixing log for a specific date"""
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # Find production run for this date
    production_run = ProductionRun.query.filter_by(date=target_date).first()
    if not production_run:
        return jsonify({'error': f'No production run found for date {date_str}'}), 404

    # Find mixing log for this production run
    mixing_log = MixingLog.query.filter_by(production_run_id=production_run.id).first()
    if not mixing_log:
        return jsonify({'error': f'No mixing log found for date {date_str}'}), 404

    # Build response with entries
    entries = []
    for entry in sorted(mixing_log.entries, key=lambda e: e.order):
        # Get DDT target for validation
        ddt_target = DDTTarget.query.filter_by(bread_name=entry.bread_name, is_active=True).first()

        temp_warnings = {}
        if entry.final_dough_temp and ddt_target:
            temp_warnings['final_dough_temp'] = {
                'in_range': ddt_target.target_temp_min <= entry.final_dough_temp <= ddt_target.target_temp_max,
                'target_min': ddt_target.target_temp_min,
                'target_max': ddt_target.target_temp_max
            }

        entries.append({
            'id': entry.id,
            'bread_name': entry.bread_name,
            'recipe_id': entry.recipe_id,
            'batch_size': entry.batch_size,
            'quantity': entry.quantity,
            'room_temp': entry.room_temp,
            'flour_temp': entry.flour_temp,
            'preferment_temp': entry.preferment_temp,
            'friction_factor': entry.friction_factor,
            'water_temp': entry.water_temp,
            'final_dough_temp': entry.final_dough_temp,
            'bulk_fermentation_notes': entry.bulk_fermentation_notes,
            'fold_schedule': entry.fold_schedule,
            'portioning_notes': entry.portioning_notes,
            'batch_notes': entry.batch_notes,
            'temp_warnings': temp_warnings
        })

    return jsonify({
        'id': mixing_log.id,
        'date': mixing_log.date.strftime('%Y-%m-%d'),
        'mixer_initials': mixing_log.mixer_initials,
        'notes': mixing_log.notes,
        'production_run_id': production_run.id,
        'batch_id': production_run.batch_id,
        'created_at': mixing_log.created_at.isoformat(),
        'entries': entries
    })


@app.route('/api/mixing-log/save', methods=['POST'])
def save_mixing_log():
    """Save or update mixing log with entries"""
    data = request.json

    # Validate required fields
    if not data.get('date'):
        return jsonify({'error': 'Date is required'}), 400

    if not data.get('mixer_initials'):
        return jsonify({'error': 'Mixer initials are required'}), 400

    if not data.get('entries') or len(data['entries']) == 0:
        return jsonify({'error': 'At least one entry is required'}), 400

    # Parse date
    try:
        log_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    # Find production run for this date
    production_run = ProductionRun.query.filter_by(date=log_date).first()
    if not production_run:
        return jsonify({'error': f'No production run found for date {log_date}'}), 404

    # Check if mixing log already exists
    existing_log = MixingLog.query.filter_by(production_run_id=production_run.id).first()

    warnings = []

    try:
        if existing_log:
            # Update existing
            mixing_log = existing_log
            mixing_log.mixer_initials = data['mixer_initials']
            mixing_log.notes = data.get('notes', '')
            mixing_log.updated_at = datetime.utcnow()

            # Delete old entries
            for entry in mixing_log.entries:
                db.session.delete(entry)
        else:
            # Create new
            mixing_log = MixingLog(
                production_run_id=production_run.id,
                date=log_date,
                mixer_initials=data['mixer_initials'],
                notes=data.get('notes', '')
            )
            db.session.add(mixing_log)

        db.session.flush()  # Get mixing_log.id

        # Add entries and validate
        for i, entry_data in enumerate(data['entries']):
            entry = MixingLogEntry(
                mixing_log_id=mixing_log.id,
                bread_name=entry_data['bread_name'],
                recipe_id=entry_data.get('recipe_id'),
                batch_size=entry_data.get('batch_size'),
                quantity=entry_data.get('quantity'),
                room_temp=entry_data.get('room_temp'),
                flour_temp=entry_data.get('flour_temp'),
                preferment_temp=entry_data.get('preferment_temp'),
                friction_factor=entry_data.get('friction_factor'),
                water_temp=entry_data.get('water_temp'),
                final_dough_temp=entry_data.get('final_dough_temp'),
                bulk_fermentation_notes=entry_data.get('bulk_fermentation_notes'),
                fold_schedule=entry_data.get('fold_schedule'),
                portioning_notes=entry_data.get('portioning_notes'),
                batch_notes=entry_data.get('batch_notes'),
                order=i
            )
            db.session.add(entry)

            # Validate against DDT targets
            if entry.final_dough_temp:
                target = DDTTarget.query.filter_by(bread_name=entry.bread_name, is_active=True).first()

                if target:
                    in_range = (target.target_temp_min <= entry.final_dough_temp <= target.target_temp_max)
                    warnings.append({
                        'bread_name': entry.bread_name,
                        'field': 'final_dough_temp',
                        'value': entry.final_dough_temp,
                        'target_range': f"{target.target_temp_min}-{target.target_temp_max}°F",
                        'in_range': in_range
                    })

        db.session.commit()

        return jsonify({
            'success': True,
            'mixing_log_id': mixing_log.id,
            'warnings': warnings
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error saving mixing log: {str(e)}")
        return jsonify({'error': 'Failed to save mixing log. Please try again.'}), 500


@app.route('/api/mixing-log/history')
def get_mixing_log_history():
    """Query mixing logs with filters"""
    # Get query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    bread_name = request.args.get('bread_name')
    limit = int(request.args.get('limit', 50))
    offset = int(request.args.get('offset', 0))

    # Build query
    query = MixingLog.query

    # Date filters
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(MixingLog.date >= start_date)
        except ValueError:
            return jsonify({'error': 'Invalid start_date format'}), 400

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(MixingLog.date <= end_date)
        except ValueError:
            return jsonify({'error': 'Invalid end_date format'}), 400

    # Bread filter (need to join with entries)
    if bread_name:
        query = query.join(MixingLogEntry).filter(MixingLogEntry.bread_name == bread_name)

    # Count total
    total = query.count()

    # Get logs with pagination
    logs = query.order_by(MixingLog.date.desc()).limit(limit).offset(offset).all()

    # Format response
    log_list = []
    for log in logs:
        # Get breads mixed
        breads_mixed = list(set([entry.bread_name for entry in log.entries]))

        # Calculate average final temp
        temps = [entry.final_dough_temp for entry in log.entries if entry.final_dough_temp]
        avg_final_temp = sum(temps) / len(temps) if temps else None

        # Check if any warnings
        has_warnings = False
        for entry in log.entries:
            if entry.final_dough_temp:
                target = DDTTarget.query.filter_by(bread_name=entry.bread_name, is_active=True).first()
                if target and not (target.target_temp_min <= entry.final_dough_temp <= target.target_temp_max):
                    has_warnings = True
                    break

        log_list.append({
            'id': log.id,
            'date': log.date.strftime('%Y-%m-%d'),
            'mixer_initials': log.mixer_initials,
            'batch_id': log.production_run.batch_id,
            'breads_mixed': breads_mixed,
            'entry_count': len(log.entries),
            'avg_final_temp': round(avg_final_temp, 1) if avg_final_temp else None,
            'has_warnings': has_warnings
        })

    return jsonify({
        'total': total,
        'logs': log_list
    })


@app.route('/api/mixing-log/trends/<bread_name>')
def get_mixing_log_trends(bread_name):
    """Get temperature trend data for a specific bread type"""
    # Get query parameters
    days = int(request.args.get('days', 30))
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    # Calculate date range
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400
    else:
        end_date = date.today()
        start_date = end_date - timedelta(days=days)

    # Get DDT target for this bread
    ddt_target = DDTTarget.query.filter_by(bread_name=bread_name, is_active=True).first()

    # Get all entries for this bread in date range
    entries = MixingLogEntry.query\
        .join(MixingLog)\
        .filter(MixingLogEntry.bread_name == bread_name)\
        .filter(MixingLog.date >= start_date)\
        .filter(MixingLog.date <= end_date)\
        .order_by(MixingLog.date.desc())\
        .all()

    # Build data points
    data_points = []
    final_temps = []
    in_range_count = 0

    for entry in entries:
        in_range = False
        if entry.final_dough_temp and ddt_target:
            in_range = (ddt_target.target_temp_min <= entry.final_dough_temp <= ddt_target.target_temp_max)
            if in_range:
                in_range_count += 1
            final_temps.append(entry.final_dough_temp)

        data_points.append({
            'date': entry.mixing_log.date.strftime('%Y-%m-%d'),
            'mixing_log_id': entry.mixing_log_id,
            'room_temp': entry.room_temp,
            'flour_temp': entry.flour_temp,
            'preferment_temp': entry.preferment_temp,
            'water_temp': entry.water_temp,
            'final_dough_temp': entry.final_dough_temp,
            'batch_size': entry.batch_size,
            'in_range': in_range
        })

    # Calculate statistics
    statistics = {}
    if final_temps:
        statistics = {
            'avg_final_temp': round(sum(final_temps) / len(final_temps), 1),
            'min_final_temp': round(min(final_temps), 1),
            'max_final_temp': round(max(final_temps), 1),
            'in_range_percentage': round((in_range_count / len(final_temps)) * 100, 1) if ddt_target else None,
            'total_entries': len(final_temps)
        }

    return jsonify({
        'bread_name': bread_name,
        'target_range': {
            'min': ddt_target.target_temp_min if ddt_target else None,
            'max': ddt_target.target_temp_max if ddt_target else None
        } if ddt_target else None,
        'data_points': data_points,
        'statistics': statistics
    })


@app.route('/issues')
def issues_page():
    """Production issues page"""
    return render_template('issues.html')


@app.route('/api/issues', methods=['GET'])
def get_issues():
    """Get all production issues"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    issue_type = request.args.get('issue_type')
    severity = request.args.get('severity')

    query = ProductionIssue.query

    if start_date:
        query = query.filter(ProductionIssue.date >= start_date)
    if end_date:
        query = query.filter(ProductionIssue.date <= end_date)
    if issue_type:
        query = query.filter(ProductionIssue.issue_type == issue_type)
    if severity:
        query = query.filter(ProductionIssue.severity == severity)

    issues = query.order_by(ProductionIssue.date.desc()).all()

    return jsonify([{
        'id': issue.id,
        'date': issue.date.strftime('%Y-%m-%d'),
        'issue_type': issue.issue_type,
        'severity': issue.severity,
        'title': issue.title,
        'description': issue.description,
        'affected_items': issue.affected_items,
        'resolution': issue.resolution,
        'resolved_at': issue.resolved_at.isoformat() if issue.resolved_at else None,
        'reported_by': issue.reported_by,
        'created_at': issue.created_at.isoformat()
    } for issue in issues])


@app.route('/api/issues', methods=['POST'])
def create_issue():
    """Create a new production issue"""
    data = request.json

    issue = ProductionIssue(
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        issue_type=data['issue_type'],
        severity=data['severity'],
        title=data['title'],
        description=data['description'],
        affected_items=data.get('affected_items', ''),
        reported_by=data.get('reported_by', '')
    )

    db.session.add(issue)
    db.session.commit()

    return jsonify({'success': True, 'issue_id': issue.id})


@app.route('/api/issues/<int:issue_id>', methods=['PUT'])
def update_issue(issue_id):
    """Update a production issue (mark as resolved, etc.)"""
    issue = ProductionIssue.query.get_or_404(issue_id)
    data = request.json

    if 'resolution' in data:
        issue.resolution = data['resolution']
        if data.get('mark_resolved'):
            issue.resolved_at = datetime.utcnow()

    if 'severity' in data:
        issue.severity = data['severity']

    db.session.commit()

    return jsonify({'success': True})


# ===== Inventory Management Endpoints =====

@app.route('/inventory')
def inventory_page():
    """Inventory management page"""
    return render_template('inventory.html')


@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    """Get all ingredients with inventory levels"""
    ingredients = Ingredient.query.all()
    inventory_data = []

    for ing in ingredients:
        # Check if low stock
        is_low_stock = False
        if ing.low_stock_threshold and ing.quantity_in_stock:
            is_low_stock = ing.quantity_in_stock <= ing.low_stock_threshold

        inventory_data.append({
            'id': ing.id,
            'name': ing.name,
            'category': ing.category,
            'unit': ing.unit,
            'quantity_in_stock': ing.quantity_in_stock or 0,
            'low_stock_threshold': ing.low_stock_threshold,
            'is_low_stock': is_low_stock,
            'cost_per_unit': ing.cost_per_unit,
            'last_updated': ing.last_updated.isoformat() if ing.last_updated else None
        })

    return jsonify(inventory_data)


@app.route('/api/inventory/add', methods=['POST'])
def add_inventory():
    """Add stock to an ingredient"""
    data = request.json
    ingredient_id = data['ingredient_id']
    quantity = float(data['quantity'])
    notes = data.get('notes', '')
    created_by = data.get('created_by', '')

    ingredient = Ingredient.query.get_or_404(ingredient_id)

    quantity_before = ingredient.quantity_in_stock or 0
    quantity_after = quantity_before + quantity

    # Update ingredient stock
    ingredient.quantity_in_stock = quantity_after
    ingredient.last_updated = datetime.utcnow()

    # Log transaction
    transaction = InventoryTransaction(
        ingredient_id=ingredient_id,
        transaction_type='addition',
        quantity=quantity,
        quantity_before=quantity_before,
        quantity_after=quantity_after,
        notes=notes,
        created_by=created_by
    )
    db.session.add(transaction)
    db.session.commit()

    return jsonify({
        'success': True,
        'new_quantity': quantity_after,
        'transaction_id': transaction.id
    })


@app.route('/api/inventory/adjust', methods=['POST'])
def adjust_inventory():
    """Adjust stock level (manual correction)"""
    data = request.json
    ingredient_id = data['ingredient_id']
    new_quantity = float(data['new_quantity'])
    notes = data.get('notes', '')
    created_by = data.get('created_by', '')

    ingredient = Ingredient.query.get_or_404(ingredient_id)

    quantity_before = ingredient.quantity_in_stock or 0
    quantity_change = new_quantity - quantity_before

    # Update ingredient stock
    ingredient.quantity_in_stock = new_quantity
    ingredient.last_updated = datetime.utcnow()

    # Log transaction
    transaction = InventoryTransaction(
        ingredient_id=ingredient_id,
        transaction_type='adjustment',
        quantity=quantity_change,
        quantity_before=quantity_before,
        quantity_after=new_quantity,
        notes=notes,
        created_by=created_by
    )
    db.session.add(transaction)
    db.session.commit()

    return jsonify({
        'success': True,
        'new_quantity': new_quantity,
        'transaction_id': transaction.id
    })


@app.route('/api/inventory/set-threshold', methods=['POST'])
def set_low_stock_threshold():
    """Set low stock threshold for an ingredient"""
    data = request.json
    ingredient_id = data['ingredient_id']
    threshold = float(data['threshold']) if data.get('threshold') else None

    ingredient = Ingredient.query.get_or_404(ingredient_id)
    ingredient.low_stock_threshold = threshold
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/inventory/transactions', methods=['GET'])
def get_inventory_transactions():
    """Get inventory transaction history"""
    ingredient_id = request.args.get('ingredient_id')
    limit = int(request.args.get('limit', 50))

    query = InventoryTransaction.query

    if ingredient_id:
        query = query.filter_by(ingredient_id=ingredient_id)

    transactions = query.order_by(InventoryTransaction.created_at.desc()).limit(limit).all()

    return jsonify([{
        'id': t.id,
        'ingredient_id': t.ingredient_id,
        'ingredient_name': t.ingredient.name if t.ingredient else 'Unknown',
        'transaction_type': t.transaction_type,
        'quantity': t.quantity,
        'quantity_before': t.quantity_before,
        'quantity_after': t.quantity_after,
        'production_run_id': t.production_run_id,
        'notes': t.notes,
        'created_by': t.created_by,
        'created_at': t.created_at.isoformat() if t.created_at else None
    } for t in transactions])


@app.route('/api/inventory/low-stock', methods=['GET'])
def get_low_stock_items():
    """Get ingredients that are below their low stock threshold"""
    low_stock_items = []

    ingredients = Ingredient.query.filter(
        Ingredient.low_stock_threshold.isnot(None)
    ).all()

    for ing in ingredients:
        if ing.quantity_in_stock and ing.quantity_in_stock <= ing.low_stock_threshold:
            low_stock_items.append({
                'id': ing.id,
                'name': ing.name,
                'category': ing.category,
                'quantity_in_stock': ing.quantity_in_stock,
                'low_stock_threshold': ing.low_stock_threshold,
                'deficit': ing.low_stock_threshold - ing.quantity_in_stock
            })

    return jsonify(low_stock_items)


# ===== Customer Production View Endpoints =====

@app.route('/customer-production')
def customer_production_page():
    """Customer production view page"""
    return render_template('customer_production.html')


@app.route('/api/customer-production/<int:customer_id>', methods=['GET'])
def get_customer_production(customer_id):
    """Get customer production calendar for a date range"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return jsonify({'error': 'start_date and end_date required'}), 400

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Get all orders for this customer in the date range
    orders = Order.query.filter(
        Order.customer_id == customer_id,
        Order.order_date >= start_date,
        Order.order_date <= end_date
    ).all()

    # Get customer info
    customer = Customer.query.get_or_404(customer_id)

    # Organize orders by recipe and date
    production_data = {}
    all_dates = []
    current_date = start_date
    while current_date <= end_date:
        all_dates.append(current_date)
        current_date += timedelta(days=1)

    # Group orders by recipe
    for order in orders:
        recipe_name = order.recipe.name
        if recipe_name not in production_data:
            production_data[recipe_name] = {
                'recipe_id': order.recipe_id,
                'dates': {}
            }

        date_str = order.order_date.strftime('%Y-%m-%d')
        production_data[recipe_name]['dates'][date_str] = order.quantity

    # Build response
    result = {
        'customer_id': customer.id,
        'customer_name': customer.name,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'dates': [d.strftime('%Y-%m-%d') for d in all_dates],
        'recipes': []
    }

    # Add each recipe with quantities for each date
    for recipe_name, data in sorted(production_data.items()):
        recipe_row = {
            'recipe_name': recipe_name,
            'recipe_id': data['recipe_id'],
            'quantities': []
        }

        for date_obj in all_dates:
            date_str = date_obj.strftime('%Y-%m-%d')
            quantity = data['dates'].get(date_str, 0)
            recipe_row['quantities'].append({
                'date': date_str,
                'day_of_week': date_obj.strftime('%A'),
                'quantity': quantity
            })

        result['recipes'].append(recipe_row)

    return jsonify(result)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host='0.0.0.0', port=port)
